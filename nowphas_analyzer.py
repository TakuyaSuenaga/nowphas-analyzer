#!/usr/bin/env python3
"""
nowphas_analyzer
"""

__author__ = "Author"
__version__ = "0.1.0"
__license__ = "MIT"

import os
import io
import glob
import datetime
import argparse
import numpy as np
import pandas as pd
import openpyxl as px
from openpyxl.chart import Reference, RadarChart


def read_file(filepath):
    """
    Read NOWPHAS data file into data frame
    """
    columns = ["年月日時", "ﾌﾗｸﾞ", "波数", "平均波波高", "平均波周期", "有義波波高", "有義波周期",
               "1/10波波高", "1/10波周期", "最高波波高", "最高波周期", "波向"]
    offset = 12 if "e" in os.path.basename(filepath) else 10
    buffer = ""
    with open(filepath, mode='r', encoding='shift-jis') as f:
        for line in f:
            buffer += line[:offset].replace(' ', '0') + line[offset:]
    df = pd.read_csv(io.StringIO(buffer),
                     delim_whitespace=True,
                     header=None,
                     skiprows=1,
                     names=columns)
    return df


def read_dir(dirpath):
    """
    All Nowphas data files in a folder are read into a data frame,
    concatenated, and reordered by datetime.
    """
    filepaths = glob.glob(os.path.join(dirpath, "[hH]*.txt"))
    df = pd.concat([read_file(filepath) for filepath in filepaths],
                   ignore_index=True)
    df = df.sort_values("年月日時")
    df = df.reset_index(drop=True)
    return df


def map_datetime(_datetime):
    """
    Convert year, month, date, and time to datetime type
    """
    if len(str(_datetime)) == 10:
        return datetime.datetime.strptime(str(_datetime), "%Y%m%d%H")
    else:
        return datetime.datetime.strptime(str(_datetime), "%Y%m%d%H%M")


def add_rader_chart(filepath, sheetname):
    """
    Create a radar chart and write it in Excel
    """
    wb = px.load_workbook(filepath)
    ws = wb[sheetname]

    # format
    from openpyxl.styles.numbers import builtin_format_code
    for index in range(3, 15):
        ws.cell(index, 3).number_format = builtin_format_code(10)
        ws.cell(index, 3).number_format = '0.0%'

    # chart
    rmin = 2
    rmax = 14
    cmin = 2
    cmax = 3

    chart = RadarChart()
    chart.type = 'standard'  # 'filled', 'marker', 'standard'
    labels = Reference(ws, min_col=cmin, min_row=rmin+1, max_row=rmax)
    data = Reference(ws, min_col=cmin+1, max_col=cmax,
                     min_row=rmin, max_row=rmax)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(labels)
    chart.legend.legendPos = "t"
    chart.title = 'Wave Direction'
    chart.height = 10
    chart.width = 10
    chart.anchor = 'B17'

    ws.add_chart(chart)
    wb.save(filepath)


def totaling(df: pd.DataFrame):
    """
    """
    # binning
    bins = [0, 15, 45, 75, 105, 135, 165, 195, 225, 255, 285, 315, 345, 360]
    labels = [1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 1]
    df['方角'] = pd.cut(df["波向"], bins=bins, labels=labels, ordered=False)
    # totaling
    total_df = pd.DataFrame({'波数合計': df.groupby('方角')["波数"].sum()})
    total = total_df["波数合計"].sum()
    total_df.insert(0, "割合", total_df["波数合計"].apply(lambda x: x / total))
    total_df.insert(1, "範囲1", list(range(15, 360, 30)))
    total_df.insert(2, "範囲2", [i % 360 for i in range(345, 690, 30)])
    # write excel
    filename = "nowphas_wave_frequency_distribution.xlsx"
    sheetname = "sheet1"
    with pd.ExcelWriter(filename) as writer:
        total_df.to_excel(writer, sheetname, startrow=1, startcol=1)
    add_rader_chart(filename, sheetname)


def make_dir_df(df: pd.DataFrame):
    """
    波高、周期ごとの波数の集計データ
    """
    if df.empty:
        df = pd.DataFrame([[0]*len(df.columns)], columns=list(df.columns))
    total = sum(df['波数'])
    dir_df = pd.pivot_table(
        df, values="波数", index=["Hs"], columns=["Tp"],
        # aggfunc=np.sum,
        aggfunc=lambda x: round(sum(x)/total * 100, 2) if total else 0,
        margins=True,
        margins_name='Total',
        fill_value=0,
        dropna=False)
    return dir_df


def write_cell(filepath, sheetname, row, col, value):
    """
    指定のセルに値を書き込む
    """
    wb = px.load_workbook(filepath)
    ws = wb[sheetname]
    ws.cell(row=row, column=col, value=value)
    wb.save(filepath)


def output_period(dir_no, df: pd.DataFrame):
    """
    波高、周期の集計データをExcelに出力
    """
    filename = "nowphas_wave_frequency_distribution.xlsx"
    sheetname = "sheet1"
    row_offset = (25 * (dir_no - 1))
    write_cell(filename, sheetname, 2+row_offset, 10, f"direction_{dir_no}")
    if not df.empty:
        with pd.ExcelWriter(
                    filename, mode='a', if_sheet_exists='overlay'
                ) as writer:
            df.to_excel(writer, sheetname, startrow=2+row_offset, startcol=9)


def frequency_distribution(df: pd.DataFrame):
    """
    """
    # binning
    bins = np.arange(0, 10.5, 0.5).tolist() + [999.9]
    labels = []
    for i in range(0, 10):
        labels.append("{:.1f}-{:.1f}".format(i, i + 0.5))
        labels.append("{:.1f}-{:.1f}".format(i + 0.5, i + 1))
    labels.append(">10.0")
    df['Hs'] = pd.cut(df["有義波波高"], bins=bins, labels=labels, ordered=False)
    bins = list(range(0, 16, 1)) + [999.9]
    labels = [f"{i}-{i+1}" for i in range(0, 15)] + [">15"]
    df['Tp'] = pd.cut(df["有義波周期"], bins=bins, labels=labels, ordered=False)
    for dir_no in range(1, 13):
        dir_df = make_dir_df(df[df["方角"].isin([dir_no])])
        output_period(dir_no, dir_df)


def nowphas_analyzer(dirpath: str) -> None:
    """
    """
    # Read files
    df = read_dir(dirpath)
    # remove illegal data
    df = df.replace(
        [66.66, 666.6, 6666, 77.77, 777.7, 7777, 99.99, 999.9, 9999], None)
    df = df.dropna()
    # change type of str to datetime
    df["年月日時"] = df["年月日時"].apply(map_datetime)
    # proccess and write excel
    totaling(df)
    frequency_distribution(df)


def main(args):
    """ Main entry point of the app """
    start = datetime.datetime.now()
    print(start.strftime('%Y-%m-%d %H:%M:%S.%f') + " [START] main()")
    nowphas_analyzer(args.dirpath)
    end = datetime.datetime.now()
    print(end.strftime('%Y-%m-%d %H:%M:%S.%f') + " [END] main()")
    print(f"time: {end - start}")


if __name__ == "__main__":
    """ This is executed when run from the command line """
    parser = argparse.ArgumentParser()

    # Required positional argument
    parser.add_argument(
        "dirpath",
        help="Path of the directory containing NOWPHAS data files.")

    args = parser.parse_args()
    main(args)
